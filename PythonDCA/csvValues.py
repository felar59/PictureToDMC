import openpyxl
import pandas as pd
from math import sqrt
import os

class dataChart:
    def __init__(self):
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(BASE_DIR, "DMCcharts2025.xlsx")
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.ws = self.wb.active

        self.df = pd.read_excel(file_path)

    def createRGBcol(self):   
        colors_rgb = []
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=3, max_col=3):
            cell = row[0]
            fill = cell.fill
            if fill and fill.fgColor.type == "rgb":
                argb = fill.fgColor.rgb 
                hex_rgb = argb[2:] 

                r = int(hex_rgb[0:2], 16)
                g = int(hex_rgb[2:4], 16)
                b = int(hex_rgb[4:6], 16)
                colors_rgb.append((r, g, b))
            else:
                colors_rgb.append(None)

        self.df["RGB"] = colors_rgb

    def get_datas(self):
        return self.df
    
    def findClosestColor(self, color, usedColor, colorsList):
        dist = 254
        r,g,b = color
        newcolor = (255, 0, 0)
        num = "404"
        name = "Error"

        if colorsList == []:
            for _, x in self.df.iterrows():
                r2, g2, b2 = x["RGB"]

                newDist = sqrt((r - r2)**2 + (g - g2)**2 + (b - b2)**2)

                if newDist < dist and x["Number"] not in usedColor:
                    dist = newDist
                    newcolor = x["RGB"]
                    num = x["Number"]
                    name = x["Name"]
        else:
            for elem in colorsList:
                rgb2 = self.from_Hex_to_Rgb(elem["hex"])
                r2, g2, b2 = rgb2

                newDist = sqrt((r - r2)**2 + (g - g2)**2 + (b - b2)**2)

                if newDist < dist and elem["num"] not in usedColor:
                    dist = newDist
                    newcolor = rgb2
                    num = elem["num"]
                    name = elem["name"]

        return (num, name, newcolor)
    
    def from_Hex_to_Rgb(self, hex):
        hex = hex.lstrip('#')  # supprime le #
        if len(hex) != 6:
            raise ValueError("La couleur hex doit avoir 6 caractères")
        
        r = int(hex[0:2], 16)
        g = int(hex[2:4], 16)
        b = int(hex[4:6], 16)
        
        return (r, g, b)

    def findNewColor(self, colors, color):

        dist = 254
        r, g, b = self.from_Hex_to_Rgb(color["hex"])
        used_numbers = [str(c["num"]) for c in colors]

        top3 = []
        for _, x in self.df.iterrows():
            num_str = str(x["Number"])
            if num_str in used_numbers or num_str == str(color["num"]):
                continue

            r2, g2, b2 = x["RGB"]
            dist = sqrt((r - r2)**2 + (g - g2)**2 + (b - b2)**2)

            top3.append((dist, x["Number"], x["Name"], x["RGB"]))
            top3.sort(key=lambda t: t[0])
            if len(top3) > 3:
                top3.pop()

        return [(num, name, rgb) for _, num, name, rgb in top3]

    def addColor(self, colorNum):
        print(self.df)
        for _, x in self.df.iterrows():
            if str(colorNum) == str(x["Number"]):
                print(str(x["Number"]), x["Name"], x["Color"])
                return (str(x["Number"]), x["Name"], x["RGB"])
    
        return ("404","Error: Not Found",(255, 0, 0))
